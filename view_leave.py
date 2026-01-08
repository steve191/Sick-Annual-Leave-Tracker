import os
import sys
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from database import collect_data_view


def rates_file():
	if getattr(sys, 'frozen', False):
		base_path = sys._MEIPASS
	else:
		base_path = os.path.dirname(os.path.abspath(__file__))

	file = os.path.join(base_path, 'rates.xlsx')

	if os.path.exists(file):
		os.startfile(file)
	else:
		# Create rates excel file
		wb = Workbook()
		ws = wb.active
		ws.title = "Rates"
		wb.save(file)
		os.startfile(file)

def get_save_path(filename):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, filename)

def view_all_leave():
	# Collect data
	records = collect_data_view()

	# Format setup
	hd_format = NamedStyle(name="heading_format")
	hd_format.font = Font(bold=True, underline='single')

	headings = {'A': 14.50, 'B': 20.78, 'C': 20.78, 'D': 10.60}

	# Start woorkbook
	wb = Workbook()

	# Delete first sheet
	del wb['Sheet']

	# Add style
	wb.add_named_style(hd_format)

	# Loop through employee info and create wookbook
	for id, rec in records.items():
		ws = wb.create_sheet(rec['info'][0])

		# Start row and col
		i_row = 0
		al_row = 0
		sl_row = 0

		# EMPLOYEE INFO
		# Add heading
		ws['A1'] = 'ID'
		ws['B1'] = 'Name'
		ws['C1'] = 'Surname'
		ws['D1'] = 'Start Date'

		# Add info
		ws['A2'] = id
		ws['B2'] = rec['info'][0]
		ws['C2'] = rec['info'][1]
		ws['D2'] = rec['info'][2]

		# Employee format
		for col, size in headings.items():
			ws.column_dimensions[col].width = size
			ws[f'{col}1'].style = 'heading_format'

		# check leave
		check_annual = 'annual' in rec
		check_sick = 'sick' in rec

		# ANNUAL LEAVE
		if check_annual == True:
			# Add annual leave headers
			ws[f'A{4}'] = 'ANNUAL LEAVE'
			ws[f'A{5}'] = 'ID'
			ws[f'B{5}'] = 'Number Of Days'
			ws[f'C{5}'] = 'Start Date'
			ws[f'D{5}'] = 'End Date'

			# Add employee annual leave
			for al in rec['annual']:
				ws[f'A{6 + al_row}'] = id
				ws[f'B{6 + al_row}'] = al[0]
				ws[f'C{6 + al_row}'] = al[1]
				ws[f'D{6 + al_row}'] = al[2]

				al_row += 1

		# Annual leave format
		ws['A4'].font = Font(bold=True)

		for col in headings.keys():
			ws[f'{col}5'].style = 'heading_format'

		# SICK LEAVE
		# Get lenght of annual leave list
		if check_annual == True:
			al_len = 6 + len(rec['annual'])

		# Add sick leave headers
		if check_sick == True:
			ws[f'A{al_len + 1}'] = 'SICK LEAVE'
			ws[f'A{al_len + 2}'] = 'ID'
			ws[f'B{al_len + 2}'] = 'Number Of Days'
			ws[f'C{al_len + 2}'] = 'Start Date'
			ws[f'D{al_len + 2}'] = 'End Date'

			# Add employee annual leave
			for sl in rec['sick']:
				ws[f'A{al_len + 3 + sl_row}'] = id
				ws[f'B{al_len + 3 + sl_row}'] = sl[0]
				ws[f'C{al_len + 3 + sl_row}'] = sl[1]
				ws[f'D{al_len + 3 + sl_row}'] = sl[2]

				sl_row += 1

			# Sick leave format
			ws[f'A{al_len + 1}'].font = Font(bold=True)

			for col in headings.keys():
				ws[f'{col}{al_len + 2}'].style = 'heading_format'

			# Align all coloumns
			for col in headings.keys():
				for cell in ws[col]:
					cell.alignment = Alignment(horizontal='center')

	# ORDER SHEETS
	# Get all sheet names and sort them alphabetically
	sorted_sheet_names = sorted(wb.sheetnames)

	# Loop through the sorted names and move each sheet to the correct position
	for index, sheet_name in enumerate(sorted_sheet_names):
		# Find the sheet's current index (position)
		ws = wb[sheet_name]
		current_index = wb.index(ws)
		
		# Calculate how far to move it (Target Position - Current Position)
		move_distance = index - current_index
		
		# Move the sheet
		wb.move_sheet(ws, offset=move_distance)

	# Save the file
	save_path = get_save_path('allLeave.xlsx')

	try:
		wb.save(save_path)
		os.startfile(save_path)
	except Exception as error:
		messagebox.showerror(title='View Leave Error', message=error)

# ###########################################################################################
# EVERYTHING ON ONE SHEETM(???)
# ###########################################################################################

# # Start row and col
# i_row = 0
# # Annual information row
# air = 0
# # Sick information row
# sir = 0

# for id, rec in records.items():
# 	# Employee info
# 	ws[f'A{1 + i_row}'] = 'ID'
# 	ws[f'B{1 + i_row}'] = 'Name'
# 	ws[f'C{1 + i_row}'] = 'Surname'
# 	ws[f'D{1 + i_row}'] = 'Start Date'

# 	ws[f'A{2 + i_row}'] = id
# 	ws[f'B{2 + i_row}'] = rec['info'][0]
# 	ws[f'C{2 + i_row}'] = rec['info'][1]
# 	ws[f'D{2 + i_row}'] = rec['info'][2]

# 	# ANNUAL LEAVE
# 	# Get max row
# 	lmxr = ws.max_row

# 	# Add annual leave headers
# 	ws[f'A{lmxr + 2}'] = 'Annual Leave'
# 	ws[f'A{lmxr + 3}'] = 'ID'
# 	ws[f'B{lmxr + 3}'] = 'Number Of Days'
# 	ws[f'C{lmxr + 3}'] = 'Start Date'
# 	ws[f'D{lmxr + 3}'] = 'End Date'

# 	i_row += lmxr
	
# 	# Add employee annual leave
# 	ai_row = ws.max_row + 1

# 	for al in rec['annual']:
# 		ws[f'A{ai_row + air}'] = id
# 		ws[f'B{ai_row + air}'] = al[0]
# 		ws[f'C{ai_row + air}'] = al[1]
# 		ws[f'D{ai_row + air}'] = al[2]

# 		air += 1
# 		i_row += 1

# 	# SICK LEAVE
# 	# Get max row
# 	smxr = ws.max_row

# 	# Add sick leave headers
# 	ws[f'A{smxr + 2}'] = 'Sick Leave'
# 	ws[f'A{smxr + 3}'] = 'ID'
# 	ws[f'B{smxr + 3}'] = 'Number Of Days'
# 	ws[f'C{smxr + 3}'] = 'Start Date'
# 	ws[f'D{smxr + 3}'] = 'End Date'

# 	i_row += smxr

# 	# Add employee annual leave
# 	si_row = ws.max_row + 1

# 	for sl in rec['sick']:
# 		ws[f'A{si_row + sir}'] = id
# 		ws[f'B{si_row + sir}'] = sl[0]
# 		ws[f'C{si_row + sir}'] = sl[1]
# 		ws[f'D{si_row + sir}'] = sl[2]

# 		sir+= 1
# 		i_row += 1