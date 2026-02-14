import os
import io
import shutil
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_wtf.csrf import CSRFProtect
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font
import db

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', os.urandom(32))
csrf = CSRFProtect(app)

def safe_sheet_name(name, emp_id, used_names):
    base = f"{name} ({emp_id})"[:31]
    if base in used_names:
        counter = 2
        while f"{base[:28]}_{counter}" in used_names:
            counter += 1
        base = f"{base[:28]}_{counter}"
    used_names.add(base)
    return base

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
BACKUP_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

@app.after_request
def add_no_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    return response

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        user = db.get_user(session['username'])
        if user and user['force_password_change'] and request.endpoint != 'change_password':
            return redirect(url_for('change_password'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.verify_password(username, password)
        if user:
            session['username'] = username
            if user['force_password_change']:
                return redirect(url_for('change_password'))
            return redirect(url_for('index'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if 'username' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'error')
        elif new_password != confirm_password:
            flash('Passwords do not match.', 'error')
        else:
            db.update_password(session['username'], new_password)
            flash('Password changed successfully.', 'success')
            return redirect(url_for('index'))
    return render_template('change_password.html')

@app.route('/')
@login_required
def index():
    employees = db.get_employee_summary()
    return render_template('index.html', employees=employees)

@app.route('/employees/add', methods=['POST'])
@login_required
def add_employee():
    try:
        db.add_employee(
            request.form['id'],
            request.form['firstName'],
            request.form['lastName'],
            request.form['startDate']
        )
        med_folder = os.path.join(UPLOAD_FOLDER, f"{request.form['firstName'].capitalize()} {request.form['lastName'].capitalize()}")
        os.makedirs(med_folder, exist_ok=True)
        flash('Employee added successfully.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('index'))

@app.route('/employees/update', methods=['POST'])
@login_required
def update_employee():
    try:
        db.update_employee(
            request.form['id'],
            request.form['firstName'],
            request.form['lastName'],
            request.form['startDate']
        )
        flash('Employee updated successfully.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('index'))

@app.route('/employees/delete', methods=['POST'])
@login_required
def delete_employee():
    emp_id = request.form['id']
    fname = request.form.get('firstName', '').strip()
    sname = request.form.get('lastName', '').strip()
    try:
        backup_employee(emp_id, fname, sname)
        db.delete_employee(emp_id)
        flash('Employee deleted and backed up.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('index'))

@app.route('/annual-leave')
@login_required
def annual_leave():
    records = db.get_annual_leave()
    return render_template('annual_leave.html', records=records)

@app.route('/annual-leave/add', methods=['GET', 'POST'])
@login_required
def add_annual_leave():
    if request.method == 'POST':
        try:
            db.add_annual_leave(
                request.form['id'],
                request.form['firstName'],
                int(request.form['days']),
                request.form['startDate'],
                request.form['endDate'],
                request.form.get('comment', '')
            )
            flash('Annual leave added.', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(str(e), 'error')
    emp_id = request.args.get('id', '')
    employee = db.get_employee(emp_id) if emp_id else None
    return render_template('add_leave.html', employee=employee, leave_type='Annual')

@app.route('/annual-leave/update', methods=['POST'])
@login_required
def update_annual_leave():
    try:
        db.update_annual_leave(
            request.form['rowid'],
            request.form['days'],
            request.form['startDate'],
            request.form['endDate'],
            request.form.get('comment', '')
        )
        flash('Annual leave updated.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('annual_leave'))

@app.route('/annual-leave/delete', methods=['POST'])
@login_required
def delete_annual_leave():
    try:
        db.delete_annual_leave(request.form['rowid'])
        flash('Annual leave record deleted.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('annual_leave'))

@app.route('/sick-leave')
@login_required
def sick_leave():
    records = db.get_sick_leave()
    return render_template('sick_leave.html', records=records)

@app.route('/sick-leave/add', methods=['GET', 'POST'])
@login_required
def add_sick_leave():
    if request.method == 'POST':
        try:
            db.add_sick_leave(
                request.form['id'],
                request.form['firstName'],
                int(request.form['days']),
                request.form['startDate'],
                request.form['endDate'],
                request.form.get('comment', '')
            )
            flash('Sick leave added.', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(str(e), 'error')
    emp_id = request.args.get('id', '')
    employee = db.get_employee(emp_id) if emp_id else None
    return render_template('add_leave.html', employee=employee, leave_type='Sick')

@app.route('/sick-leave/update', methods=['POST'])
@login_required
def update_sick_leave():
    try:
        db.update_sick_leave(
            request.form['rowid'],
            request.form['days'],
            request.form['startDate'],
            request.form['endDate'],
            request.form.get('comment', '')
        )
        flash('Sick leave updated.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('sick_leave'))

@app.route('/sick-leave/delete', methods=['POST'])
@login_required
def delete_sick_leave():
    try:
        db.delete_sick_leave(request.form['rowid'])
        flash('Sick leave record deleted.', 'success')
    except Exception as e:
        flash(str(e), 'error')
    return redirect(url_for('sick_leave'))

@app.route('/export/all-leave')
@login_required
def export_all_leave():
    records = db.get_all_leave_data()

    hd_format = NamedStyle(name="heading_format")
    hd_format.font = Font(bold=True, underline='single')

    headings = {'A': 14.50, 'B': 20.78, 'C': 20.78, 'D': 10.60, 'E': 125.78}

    wb = Workbook()
    del wb['Sheet']
    wb.add_named_style(hd_format)
    used_names = set()

    for emp_id, rec in records.items():
        sheet_name = safe_sheet_name(rec['info'][0], emp_id, used_names)
        ws = wb.create_sheet(sheet_name)

        al_row = 0
        sl_row = 0

        ws['A1'] = 'ID'
        ws['B1'] = 'Name'
        ws['C1'] = 'Surname'
        ws['D1'] = 'Start Date'

        ws['A2'] = emp_id
        ws['B2'] = rec['info'][0]
        ws['C2'] = rec['info'][1]
        ws['D2'] = rec['info'][2]

        for col, size in headings.items():
            ws.column_dimensions[col].width = size
            ws[f'{col}1'].style = 'heading_format'

        if rec['annual']:
            ws['A4'] = 'ANNUAL LEAVE'
            ws['A5'] = 'ID'
            ws['B5'] = 'Number Of Days'
            ws['C5'] = 'Start Date'
            ws['D5'] = 'End Date'
            ws['E5'] = 'Comments'

            for al in rec['annual']:
                ws[f'A{6 + al_row}'] = emp_id
                ws[f'B{6 + al_row}'] = al[0]
                ws[f'C{6 + al_row}'] = al[1]
                ws[f'D{6 + al_row}'] = al[2]
                ws[f'E{6 + al_row}'] = al[3]
                al_row += 1

            ws['A4'].font = Font(bold=True)
            for col in headings.keys():
                ws[f'{col}5'].style = 'heading_format'

        al_len = 6 + len(rec['annual']) if rec['annual'] else 4

        if rec['sick']:
            ws[f'A{al_len + 1}'] = 'SICK LEAVE'
            ws[f'A{al_len + 2}'] = 'ID'
            ws[f'B{al_len + 2}'] = 'Number Of Days'
            ws[f'C{al_len + 2}'] = 'Start Date'
            ws[f'D{al_len + 2}'] = 'End Date'
            ws[f'E{al_len + 2}'] = 'Comments'

            for sl in rec['sick']:
                ws[f'A{al_len + 3 + sl_row}'] = emp_id
                ws[f'B{al_len + 3 + sl_row}'] = sl[0]
                ws[f'C{al_len + 3 + sl_row}'] = sl[1]
                ws[f'D{al_len + 3 + sl_row}'] = sl[2]
                ws[f'E{al_len + 3 + sl_row}'] = sl[3]
                sl_row += 1

            ws[f'A{al_len + 1}'].font = Font(bold=True)
            for col in headings.keys():
                ws[f'{col}{al_len + 2}'].style = 'heading_format'

        for col in headings.keys():
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal='center')

    sorted_names = sorted(wb.sheetnames)
    for i, name in enumerate(sorted_names):
        ws = wb[name]
        current = wb.index(ws)
        wb.move_sheet(ws, offset=i - current)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='allLeave.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/documents')
@login_required
def documents():
    folders = []
    if os.path.exists(UPLOAD_FOLDER):
        folders = sorted([d for d in os.listdir(UPLOAD_FOLDER)
                         if os.path.isdir(os.path.join(UPLOAD_FOLDER, d))])
    return render_template('documents.html', folders=folders)

@app.route('/documents/<folder>')
@login_required
def view_folder(folder):
    folder_path = os.path.join(UPLOAD_FOLDER, secure_filename(folder))
    files = []
    if os.path.exists(folder_path):
        files = sorted(os.listdir(folder_path))
    return render_template('folder.html', folder=folder, files=files)

@app.route('/documents/<folder>/upload', methods=['POST'])
@login_required
def upload_document(folder):
    folder_path = os.path.join(UPLOAD_FOLDER, secure_filename(folder))
    os.makedirs(folder_path, exist_ok=True)
    if 'file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('view_folder', folder=folder))
    file = request.files['file']
    if file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('view_folder', folder=folder))
    filename = secure_filename(file.filename)
    file.save(os.path.join(folder_path, filename))
    flash(f'File "{filename}" uploaded.', 'success')
    return redirect(url_for('view_folder', folder=folder))

@app.route('/documents/<folder>/<filename>/download')
@login_required
def download_document(folder, filename):
    folder_path = os.path.join(UPLOAD_FOLDER, secure_filename(folder))
    return send_file(os.path.join(folder_path, secure_filename(filename)), as_attachment=True)

@app.route('/documents/<folder>/<filename>/delete', methods=['POST'])
@login_required
def delete_document(folder, filename):
    folder_path = os.path.join(UPLOAD_FOLDER, secure_filename(folder))
    file_path = os.path.join(folder_path, secure_filename(filename))
    if os.path.exists(file_path):
        os.remove(file_path)
        flash(f'File "{filename}" deleted.', 'success')
    return redirect(url_for('view_folder', folder=folder))

def backup_employee(emp_id, fname, sname):
    records = db.get_all_leave_data(emp_id)
    backup_dir = os.path.join(BACKUP_FOLDER, f"{fname} {sname}")
    os.makedirs(backup_dir, exist_ok=True)

    hd_format = NamedStyle(name="heading_format")
    hd_format.font = Font(bold=True, underline='single')
    headings = {'A': 14.50, 'B': 20.78, 'C': 20.78, 'D': 10.60}

    wb = Workbook()
    del wb['Sheet']
    wb.add_named_style(hd_format)

    for eid, rec in records.items():
        ws = wb.create_sheet(rec['info'][0])
        al_row = 0
        sl_row = 0

        ws['A1'] = 'ID'
        ws['B1'] = 'Name'
        ws['C1'] = 'Surname'
        ws['D1'] = 'Start Date'

        ws['A2'] = eid
        ws['B2'] = rec['info'][0]
        ws['C2'] = rec['info'][1]
        ws['D2'] = rec['info'][2]

        for col, size in headings.items():
            ws.column_dimensions[col].width = size
            ws[f'{col}1'].style = 'heading_format'

        if rec['annual']:
            ws['A4'] = 'ANNUAL LEAVE'
            ws['A5'] = 'ID'
            ws['B5'] = 'Number Of Days'
            ws['C5'] = 'Start Date'
            ws['D5'] = 'End Date'
            for al in rec['annual']:
                ws[f'A{6 + al_row}'] = eid
                ws[f'B{6 + al_row}'] = al[0]
                ws[f'C{6 + al_row}'] = al[1]
                ws[f'D{6 + al_row}'] = al[2]
                al_row += 1

            ws['A4'].font = Font(bold=True)
            for col in headings.keys():
                ws[f'{col}5'].style = 'heading_format'

        al_len = 6 + len(rec['annual']) if rec['annual'] else 4

        if rec['sick']:
            ws[f'A{al_len + 1}'] = 'SICK LEAVE'
            ws[f'A{al_len + 2}'] = 'ID'
            ws[f'B{al_len + 2}'] = 'Number Of Days'
            ws[f'C{al_len + 2}'] = 'Start Date'
            ws[f'D{al_len + 2}'] = 'End Date'
            for sl in rec['sick']:
                ws[f'A{al_len + 3 + sl_row}'] = eid
                ws[f'B{al_len + 3 + sl_row}'] = sl[0]
                ws[f'C{al_len + 3 + sl_row}'] = sl[1]
                ws[f'D{al_len + 3 + sl_row}'] = sl[2]
                sl_row += 1

            ws[f'A{al_len + 1}'].font = Font(bold=True)
            for col in headings.keys():
                ws[f'{col}{al_len + 2}'].style = 'heading_format'

        for col in headings.keys():
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal='center')

    wb.save(os.path.join(backup_dir, f"{fname} {sname}.xlsx"))

    src = os.path.join(UPLOAD_FOLDER, f"{fname} {sname}")
    if os.path.exists(src):
        shutil.copytree(src, backup_dir, dirs_exist_ok=True)
        shutil.rmtree(src, ignore_errors=True)

if __name__ == '__main__':
    db.init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
