import os
import shutil
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from database import collect_data_view
from database import delete_employee_db


# Make folder if not exists
try:
	if not os.path.exists('Employee Backups'):
		os.mkdir('Employee Backups')
except Exception as error:
	messagebox.showerror(title='Add Employee Backups Folder', message=error)

def backup(id, fname, sname):
	response = messagebox.askyesno(title='Delete Employee', message='Are You Sure You Want To Deleted Employee')
	
	if response == 1:
		# Create backup folder
		if not os.path.exists(f'Employee Backups/{fname} {sname}'):
			os.mkdir(f'Employee Backups/{fname} {sname}')

		# Collect data
		records = collect_data_view(id)

		# Format setup
		hd_format = NamedStyle(name="heading_format")
		hd_format.font = Font(bold=True, underline='single')

		headings = {'A': 14.50, 'B': 20.78, 'C': 20.78, 'D': 10.60}

		# Start woorkbook
		wb = Workbook()

		# Delete first sheet
		del wb['Sheet']

		# Add style
		wb.add_named_style(hd_format)

		# Loop through employee info and create wookbook
		for id, rec in records.items():
			ws = wb.create_sheet(rec['info'][0])

			# Start row and col
			i_row = 0
			al_row = 0
			sl_row = 0

			# EMPLOYEE INFO
			# Add heading
			ws['A1'] = 'ID'
			ws['B1'] = 'Name'
			ws['C1'] = 'Surname'
			ws['D1'] = 'Start Date'

			# Add info
			ws['A2'] = id
			ws['B2'] = rec['info'][0]
			ws['C2'] = rec['info'][1]
			ws['D2'] = rec['info'][2]

			# Employee format
			for col, size in headings.items():
				ws.column_dimensions[col].width = size
				ws[f'{col}1'].style = 'heading_format'

			# check leave
			check_annual = 'annual' in rec
			check_sick = 'sick' in rec

			# ANNUAL LEAVE
			if check_annual == True:
				# Add annual leave headers
				ws[f'A{4}'] = 'ANNUAL LEAVE'
				ws[f'A{5}'] = 'ID'
				ws[f'B{5}'] = 'Number Of Days'
				ws[f'C{5}'] = 'Start Date'
				ws[f'D{5}'] = 'End Date'

				# Add employee annual leave
				for al in rec['annual']:
					ws[f'A{6 + al_row}'] = id
					ws[f'B{6 + al_row}'] = al[0]
					ws[f'C{6 + al_row}'] = al[1]
					ws[f'D{6 + al_row}'] = al[2]

					al_row += 1

			# Annual leave format
			ws['A4'].font = Font(bold=True)

			for col in headings.keys():
				ws[f'{col}5'].style = 'heading_format'

			# SICK LEAVE
			# Get lenght of annual leave list
			if check_annual == True:
				al_len = 6 + len(rec['annual'])

			# Add sick leave headers
			if check_sick == True:
				ws[f'A{al_len + 1}'] = 'SICK LEAVE'
				ws[f'A{al_len + 2}'] = 'ID'
				ws[f'B{al_len + 2}'] = 'Number Of Days'
				ws[f'C{al_len + 2}'] = 'Start Date'
				ws[f'D{al_len + 2}'] = 'End Date'

				# Add employee annual leave
				for sl in rec['sick']:
					ws[f'A{al_len + 3 + sl_row}'] = id
					ws[f'B{al_len + 3 + sl_row}'] = sl[0]
					ws[f'C{al_len + 3 + sl_row}'] = sl[1]
					ws[f'D{al_len + 3 + sl_row}'] = sl[2]

					sl_row += 1

				# Sick leave format
				ws[f'A{al_len + 1}'].font = Font(bold=True)

				for col in headings.keys():
					ws[f'{col}{al_len + 2}'].style = 'heading_format'

				# Align all coloumns
				for col in headings.keys():
					for cell in ws[col]:
						cell.alignment = Alignment(horizontal='center')

		try:
			wb.save(f'Employee Backups/{fname} {sname}/{fname} {sname}.xlsx')
		except Exception as error:
			messagebox.showerror(title='View Leave Error', message=error)

		# Copy over medical docs and delete
		try:
			src = os.path.join("Medical Documents", f"{fname} {sname}")
			dst = os.path.join("Employee Backups", f"{fname} {sname}")
			shutil.copytree(src, dst, dirs_exist_ok=True)
		except Exception as error:
			messagebox.showerror(title='Copy Of Medical Documents', message=error)

		try:
			shutil.rmtree(src)
		except FileNotFoundError:
			messagebox.showerror(title='Deletion Of Medical Documents', message="Folder Does Not Exist")
		except PermissionError:
			messagebox.showerror(title='Deletion Of Medical Documents', message="Permission Denied (Folder Might Be In Use)")

		# Delete from database
		delete_employee_db(id)